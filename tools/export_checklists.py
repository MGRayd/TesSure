"""Export the app's source-controlled checklist to an editable Excel workbook."""

from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "app/src/main/java/uk/co/pactsolutions/teslachecklist/MainActivity.java"
OUTPUT = ROOT / "TesSure_Checklist_Review.xlsx"

ITEM_PATTERN = re.compile(
    r'new CheckItem\("((?:[^"\\]|\\.)*)",\s*"((?:[^"\\]|\\.)*)"\)'
)


def java_unescape(value: str) -> str:
    return (
        value.replace(r"\"", '"')
        .replace(r"\n", "\n")
        .replace(r"\t", "\t")
        .replace(r"\\", "\\")
    )


source_text = SOURCE.read_text(encoding="utf-8")
items = [(java_unescape(section), java_unescape(text)) for section, text in ITEM_PATTERN.findall(source_text)]
if not items:
    raise RuntimeError(f"No checklist items found in {SOURCE}")

wb = Workbook()
ws = wb.active
ws.title = "Checklist Review"
info = wb.create_sheet("Instructions")
lists = wb.create_sheet("Lists")

headers = [
    "Existing ID",
    "Models",
    "Current Section",
    "Current Check",
    "Action",
    "Updated Section",
    "Updated Check",
    "Reviewer Notes",
]
ws.append(headers)
for index, (section, text) in enumerate(items):
    ws.append([f"CHECK-{index + 1:03d}", "Model 3, Model Y", section, text, "Keep", "", "", ""])

# Blank rows make additions explicit without modifying or reusing an existing ID.
for _ in range(20):
    ws.append(["", "Model 3, Model Y", "", "", "Add", "", "", ""])

header_fill = PatternFill("solid", fgColor="B51F2E")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ws.max_row}"
ws.row_dimensions[1].height = 28
widths = {"A": 14, "B": 21, "C": 30, "D": 67, "E": 13, "F": 30, "G": 67, "H": 42}
for column, width in widths.items():
    ws.column_dimensions[column].width = width
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

table = Table(displayName="ChecklistReview", ref=f"A1:H{ws.max_row}")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False
)
ws.add_table(table)

actions = ["Keep", "Update", "Delete", "Add"]
lists.append(["Allowed Actions"])
for action in actions:
    lists.append([action])
lists.sheet_state = "hidden"
validation = DataValidation(type="list", formula1="'Lists'!$A$2:$A$5", allow_blank=False)
validation.error = "Choose Keep, Update, Delete, or Add."
validation.errorTitle = "Invalid action"
validation.prompt = "Choose what should happen to this checklist item."
validation.promptTitle = "Review action"
validation.showErrorMessage = True
validation.showInputMessage = True
ws.add_data_validation(validation)
validation.add(f"E2:E{ws.max_row}")

update_fill = PatternFill("solid", fgColor="FFF2CC")
delete_fill = PatternFill("solid", fgColor="F4CCCC")
add_fill = PatternFill("solid", fgColor="D9EAD3")
ws.conditional_formatting.add(
    f"A2:H{ws.max_row}", FormulaRule(formula=["$E2=\"Update\""], fill=update_fill)
)
ws.conditional_formatting.add(
    f"A2:H{ws.max_row}", FormulaRule(formula=["$E2=\"Delete\""], fill=delete_fill)
)
ws.conditional_formatting.add(
    f"A2:H{ws.max_row}", FormulaRule(formula=["$E2=\"Add\""], fill=add_fill)
)

instructions = [
    ("TesSure checklist review", "Edit the Checklist Review sheet, then return this workbook."),
    ("Scope", f"The {len(items)} built-in checks are shared by Model 3 and Model Y. Custom checks created inside the app are stored on each device and are not present in source code."),
    ("Keep", "Leave the current section and wording unchanged."),
    ("Update", "Enter the replacement section and/or wording in the Updated columns. Blank updated fields mean no change to that field."),
    ("Delete", "Mark an existing row Delete. Keep its Existing ID unchanged."),
    ("Add", "Use a green blank row. Enter Updated Section and Updated Check; leave Existing ID blank."),
    ("Ordering", "Keep rows in the desired final order. You may move rows, including Add rows."),
    ("IDs", "Do not edit Existing ID values; they map feedback back to the current source list."),
]
for row_index, (title, detail) in enumerate(instructions, start=1):
    info.cell(row_index, 1, title)
    info.cell(row_index, 2, detail)
    info.cell(row_index, 1).font = Font(bold=True, color="B51F2E")
    info.cell(row_index, 1).alignment = Alignment(vertical="top", wrap_text=True)
    info.cell(row_index, 2).alignment = Alignment(vertical="top", wrap_text=True)
info.column_dimensions["A"].width = 20
info.column_dimensions["B"].width = 100
info.freeze_panes = "A2"

wb.save(OUTPUT)

# Reopen to catch malformed workbook output and verify the row count.
check = load_workbook(OUTPUT, read_only=True, data_only=False)
review = check["Checklist Review"]
assert review.max_row == len(items) + 21
assert review["A2"].value == "CHECK-001"
check.close()
print(f"Created {OUTPUT} with {len(items)} checklist items and 20 blank addition rows.")
